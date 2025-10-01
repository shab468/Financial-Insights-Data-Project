"""
Rebuilds the Financial Market Insights Dashboard Excel file from CSVs in ./data
Creates polished sheets: RawData, Metrics, Summary, Dashboard (with image charts)
"""

import os
from datetime import datetime
from pathlib import Path

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.drawing.image import Image as XLImage

BASE_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE_DIR, "data")
OUT_XLSX = os.path.join(BASE_DIR, "Financial_Market_Insights_Dashboard.xlsx")
CHART_DIR = os.path.join(BASE_DIR, "charts")


# -----------------------------
# Load & compute
# -----------------------------
def load_data() -> pd.DataFrame:
    frames = []
    for fname in os.listdir(DATA_DIR):
        if not fname.lower().endswith(".csv"):
            continue
        fp = os.path.join(DATA_DIR, fname)
        df = pd.read_csv(fp)

        # Normalize + coerce types
        df.columns = [c.strip().title() for c in df.columns]
        if not {"Date", "Ticker", "Close"}.issubset(df.columns):
            continue
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
        df["Ticker"] = df["Ticker"].astype(str)
        df["Close"] = pd.to_numeric(df["Close"], errors="coerce")
        df = df.dropna(subset=["Date", "Close"]).sort_values(["Ticker", "Date"])
        frames.append(df[["Date", "Ticker", "Close"]])

    if not frames:
        raise RuntimeError("No CSVs with the required columns (Date, Ticker, Close) were found in ./data")

    return pd.concat(frames, ignore_index=True).sort_values(["Ticker", "Date"])


def compute_metrics(prices: pd.DataFrame) -> pd.DataFrame:
    prices = prices.copy()
    prices["Date"] = pd.to_datetime(prices["Date"])

    parts = []
    for t, g in prices.groupby("Ticker"):
        g = g.sort_values("Date").reset_index(drop=True)
        g["Close"] = pd.to_numeric(g["Close"], errors="coerce")
        g = g.dropna(subset=["Close"])
        if g.empty:
            continue
        g["Pct_Change"] = g["Close"].pct_change()
        g["MA_10"] = g["Close"].rolling(window=10).mean()
        g["MA_30"] = g["Close"].rolling(window=30).mean()
        g["Vol_10"] = g["Pct_Change"].rolling(window=10).std() * (252 ** 0.5)  # annualized approx
        parts.append(g)

    if not parts:
        return pd.DataFrame(columns=["Date", "Ticker", "Close", "Pct_Change", "MA_10", "MA_30", "Vol_10"])

    m = pd.concat(parts, ignore_index=True)
    m["Date"] = m["Date"].dt.date
    return m


def compute_summary(metrics: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for t, g in metrics.groupby("Ticker"):
        g = g.sort_values("Date").reset_index(drop=True)
        if g.empty:
            continue
        last_close = float(g["Close"].iloc[-1]) if pd.notna(g["Close"].iloc[-1]) else None
        if len(g) >= 2 and g["Close"].iloc[0] not in (None, 0):
            first_close = float(g["Close"].iloc[0])
            total_ret = (last_close / first_close) - 1 if first_close else None
        else:
            total_ret = None
        if len(g) >= 6 and pd.notna(g["Close"].iloc[-6]):
            ret_5d = (last_close / float(g["Close"].iloc[-6])) - 1
        else:
            ret_5d = None
        vol10 = g["Vol_10"].dropna().iloc[-1] if not g["Vol_10"].dropna().empty else None
        rows.append({"Ticker": t, "Total_Return": total_ret, "Return_5D": ret_5d, "Vol_10": vol10, "Last_Close": last_close})
    return pd.DataFrame(rows, columns=["Ticker", "Total_Return", "Return_5D", "Vol_10", "Last_Close"])


# -----------------------------
# Chart rendering (Matplotlib)
# -----------------------------
def render_charts(metrics: pd.DataFrame, summary: pd.DataFrame) -> dict:
    """Create PNG charts and return their file paths."""
    Path(CHART_DIR).mkdir(parents=True, exist_ok=True)
    out = {}

    # --- Price Trend (line) ---
    pvt = metrics.copy()
    pvt["Date"] = pd.to_datetime(pvt["Date"])
    pivot = pvt.pivot(index="Date", columns="Ticker", values="Close")
    plt.figure(figsize=(11, 5))
    pivot.plot()  # no explicit colors; simple & robust
    plt.title("Price Trend (Close)")
    plt.xlabel("Date")
    plt.ylabel("Price")
    plt.tight_layout()
    price_png = os.path.join(CHART_DIR, "price_trend.png")
    plt.savefig(price_png, dpi=160)
    plt.close()
    out["price"] = price_png

    # --- Total Return (bar) ---
    if not summary.empty:
        plt.figure(figsize=(9, 4))
        s = summary.set_index("Ticker")["Total_Return"].sort_values(ascending=False)
        s.plot(kind="bar")
        plt.title("Total Return (Period)")
        plt.ylabel("Return")
        ax = plt.gca()
        ax.yaxis.set_major_formatter(mtick.PercentFormatter(1.0))
        plt.tight_layout()
        bar_png = os.path.join(CHART_DIR, "total_return.png")
        plt.savefig(bar_png, dpi=160)
        plt.close()
        out["bar"] = bar_png

    return out


# -----------------------------
# Excel building (images)
# -----------------------------
def _format_as_table(ws, name: str):
    max_row, max_col = ws.max_row, ws.max_column
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)


from openpyxl.utils import get_column_letter  # (import placed here so linter doesn't complain above)


def build_workbook_with_images(raw: pd.DataFrame, metrics: pd.DataFrame, summary: pd.DataFrame, charts: dict) -> None:
    if os.path.exists(OUT_XLSX):
        os.remove(OUT_XLSX)

    wb = Workbook()

    # RawData
    ws_raw = wb.active
    ws_raw.title = "RawData"
    for r in dataframe_to_rows(raw, index=False, header=True):
        ws_raw.append(r)
    ws_raw.freeze_panes = "A2"

    # Metrics
    ws_met = wb.create_sheet("Metrics")
    for r in dataframe_to_rows(metrics, index=False, header=True):
        ws_met.append(r)
    ws_met.freeze_panes = "A2"

    # Summary
    ws_sum = wb.create_sheet("Summary")
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws_sum.append(r)
    ws_sum.freeze_panes = "A2"
    # simple % and number formats
    for r in range(2, ws_sum.max_row + 1):
        ws_sum[f"B{r}"].number_format = "0.00%"
        ws_sum[f"C{r}"].number_format = "0.00%"
        ws_sum[f"D{r}"].number_format = "0.00%"
        ws_sum[f"E{r}"].number_format = "#,##0.00"
    # color scale for Total_Return
    if ws_sum.max_row > 1:
        ws_sum.conditional_formatting.add(
            f"B2:B{ws_sum.max_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    # Make RawData/Metrics/Summary pretty tables
    _format_as_table(ws_raw, "tblRawData")
    _format_as_table(ws_met, "tblMetrics")
    _format_as_table(ws_sum, "tblSummary")

    # Dashboard
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash["A1"] = "Financial Market Insights Dashboard"
    ws_dash["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # Insert images if they exist
    row_anchor = 4
    if "price" in charts and os.path.exists(charts["price"]):
        img1 = XLImage(charts["price"])
        ws_dash.add_image(img1, f"A{row_anchor}")   # A4
        row_anchor += 22                             # leave space under first chart

    if "bar" in charts and os.path.exists(charts["bar"]):
        img2 = XLImage(charts["bar"])
        ws_dash.add_image(img2, f"A{row_anchor}")   # A26 (approx)

    wb.save(OUT_XLSX)


# -----------------------------
# Main
# -----------------------------
def main():
    raw = load_data()
    metrics = compute_metrics(raw)
    summary = compute_summary(metrics)
    charts = render_charts(metrics, summary)
    build_workbook_with_images(raw, metrics, summary, charts)
    print(f"Dashboard written to: {OUT_XLSX}")


if __name__ == "__main__":
    main()
